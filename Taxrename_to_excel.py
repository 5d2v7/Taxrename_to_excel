import pdfplumber
from openpyxl import Workbook, load_workbook
import re
import os
import glob
import configparser
import logging
import shutil

# 读取配置文件并设置路径变量
config = configparser.ConfigParser()
config.read('config.ini', encoding='utf-8')
pdf_directory_path = config['Paths']['PDFDirectoryPath']
output_directory_path = config['Paths']['OutputDirectoryPath']

# 读取税种列表
tax_types = [tax.strip() for tax in config['TaxTypes'].values()]

# 设置日志记录
log_file_path = os.path.join(output_directory_path, 'process_log.log')
logging.basicConfig(filename=log_file_path, level=logging.INFO,
                    format='%(asctime)s:%(levelname)s:%(message)s')

# 清洗数据函数：去除换行符和不必要的文本
def clean_data(row_data):
    cleaned_data = []
    for cell in row_data:
        if cell:
            # 去除换行符
            cell = cell.replace('\n', '')
            # 去除“纳税人名称：”和“金额单位：人民币元(列至角分)”
            cell = re.sub(r'^纳税人名称：|金额单位：人民币元\(列至角分\)', '', cell).strip()
        cleaned_data.append(cell if cell else '')
    return cleaned_data

# 提取信息并重命名Excel文件的函数
def extract_info_and_rename_excel(pdf_path, excel_path):
    # 初始化工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # 使用pdfplumber打开PDF并提取表格数据
    with pdfplumber.open(pdf_path) as pdf:
        # 限制提取的页数
        pages_to_extract = range(min(2, len(pdf.pages)))
        for page_number in pages_to_extract:
            page = pdf.pages[page_number]
            table = page.extract_table()
            if table:
                for row in table:
                    ws.append(clean_data(row))

    # 保存并关闭工作簿
    wb.save(excel_path)
    wb.close()

    # 重新打开工作簿以提取公司名称、税种和受理日期
    wb = load_workbook(excel_path)
    ws = wb.active

    # 初始化变量
    company_name = None
    company_patterns = ["有限公司", "有限责任公司", "合伙企业"]
    extracted_tax_types = []
    acceptance_date = None

    # 正则表达式模式
    date_pattern = r'\s*受理日期：\s*(\d{4})\s*年\s*(\d{1,2})\s*月\s*'
    full_date_pattern = r'\s*(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日\s*'

    # 遍历工作表中的单元格以提取信息
    for row in ws.iter_rows(values_only=True):
        for cell_index, cell in enumerate(row):
            if cell:
                # 提取公司名称
                if any(company_pattern in cell for company_pattern in company_patterns) and "其他有限责任公司" not in cell:
                    company_name = cell.strip()
                    break
                # 提取税种
                for tax_type in tax_types:
                    if tax_type in cell:
                        extracted_tax_types.append(tax_type)
                # 提取受理日期
                match = re.search(date_pattern, cell)
                if match:
                    acceptance_date = f"{match.group(1)}-{match.group(2).zfill(2)}"
                elif "受理日期：" in cell:
                    next_cell = row[cell_index + 1] if cell_index + 1 < len(row) else None
                    if next_cell and re.match(full_date_pattern, next_cell):
                        date_match = re.match(full_date_pattern, next_cell)
                        acceptance_date = f"{date_match.group(1)}-{date_match.group(2).zfill(2)}"

    # 如果第一页没有公司名称，则在第一页文本中搜索
    if not company_name and len(pdf.pages) > 0:
        first_page_text = pdf.pages[0].extract_text()
        for pattern in company_patterns:
            match = re.search(rf".*{pattern}.*", first_page_text)
            if match:
                # 应用清洗函数
                company_name = clean_data([match.group(0)])[0]
                break

    wb.close()
    print(f"公司名称: {company_name}, 税种: {extracted_tax_types}, 受理日期: {acceptance_date}")
    # 如果提取到足够信息，则重命名文件
    if company_name and extracted_tax_types and acceptance_date:
        new_file_name = f"{company_name}_{extracted_tax_types[0]}_{acceptance_date}.xlsx"
        new_file_path = os.path.join(output_directory_path, new_file_name)
        os.rename(excel_path, new_file_path)
        logging.info(f'文件 {new_file_path} 已被重命名。')

        # 复制和重命名PDF文件
        new_pdf_name = f"{company_name}_{extracted_tax_types[0]}_{acceptance_date}.pdf"
        new_pdf_path = os.path.join(output_directory_path, new_pdf_name)
        shutil.copyfile(pdf_path, new_pdf_path)
        logging.info(f'PDF文件 {new_pdf_path} 已被复制并重命名。')
        return new_file_path, new_pdf_path
    else:
        # 如果未提取到足够信息，则删除Excel文件并记录警告
        os.remove(excel_path)
        logging.warning(f'文件 {pdf_path} 未能提取足够信息，未进行重命名。')
        return None, None

# 处理指定路径下的所有PDF文件
def process_pdfs_in_directory(pdf_directory_path, output_directory_path):
    pdf_files = glob.glob(os.path.join(pdf_directory_path, '*.pdf'))
    for pdf_file in pdf_files:
        excel_file = os.path.join(output_directory_path, os.path.basename(pdf_file).replace('.pdf', '.xlsx'))
        new_excel_path, new_pdf_path = extract_info_and_rename_excel(pdf_file, excel_file)
        if new_excel_path and new_pdf_path:
            logging.info(f'Excel文件 {new_excel_path} 和 PDF文件 {new_pdf_path} 已重命名。')
        else:
            logging.warning(f'未能提取足够信息，未对文件 {pdf_file} 进行重命名。')

# 开始处理PDF文件
process_pdfs_in_directory(pdf_directory_path, output_directory_path)
